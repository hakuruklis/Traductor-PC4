import pprint, json
from openpyxl import Workbook
from openpyxl import load_workbook
import copy
import pprint


wb = load_workbook(filename = 'Palabras.xlsx')
ws=wb.active


diccionario=[]
palabras = {}

z=0
for z in range(1086):
	palabras['id']=z
	palabras['palabra']=''
	palabras['sininimos'] = []
	palabras['definicion']=[]
	palabra=ws['A'+str(z+1)].value
	try:
		nueva_palabra=palabra.split(". ")
		nueva_palabra[0]=nueva_palabra[0].split(" | ")
		nueva_palabra[1]=nueva_palabra[1].split(", ")
	except:
		print(str(z+1) + "No agregado")
	x=0
	y=0
	for x in range(len(nueva_palabra)):
		if x == 0: 
			for y in range(len(nueva_palabra[0])):
				if y == 0:
					palabras['palabra']=nueva_palabra[0][0]
				else:
					palabras['sininimos'].append(nueva_palabra[0][y])
				y+=1
		elif x==1:
			y=0
			for y in range(len(nueva_palabra[1])):
					palabras['definicion'].append(nueva_palabra[1][y])
	#print(palabras)
	palabrass = copy.deepcopy(palabras)
	z+=1
	diccionario.append(palabrass)
	

pprint.pprint(diccionario)
#print(palabras)


with open('palabras.json', 'w') as outfile:  
    json.dump(diccionario, outfile)










